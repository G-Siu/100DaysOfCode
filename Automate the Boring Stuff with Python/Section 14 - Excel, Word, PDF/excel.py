import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

sheet = wb["Sheet1"]
# print(type(sheet))
# print(wb.sheetnames)  # Returns sheet names
cell = sheet["A1"]
# print(cell.value)  # Gives value in cell A1

# print(sheet.cell(row=1, column=2))  # Gives the cell number

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)  # Prints values in B1:B7
