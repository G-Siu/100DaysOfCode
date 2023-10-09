import openpyxl

wb = openpyxl.Workbook()

# print(wb.sheetnames)
sheet = wb["Sheet"]
# print(sheet["A1"].value == None)

sheet["A1"] = 42  # Assigning values to cells
sheet["A2"] = "Hello"

wb.save("example2.xlsx")  # Save a new workbook

sheet2 = wb.create_sheet()  # Create new sheet
sheet2.title = "My New Sheet Name"
wb.save("example3.xlsx")  # Good practice to save into new file

wb.create_sheet(index=0, title="My Other Sheet")  # Create sheet as first sheet
wb.save("example4.xlsx")
